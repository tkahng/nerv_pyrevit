from openpyxl.styles import Border, Side, PatternFill, Font
import os, sys
import re
import csv

def Importcsv(Filename):
    flat_list = []
    with open(Filename, 'r') as f:
        reader = csv.reader(f)
        Lst = list(reader)
        for sublist in Lst:
            for item in sublist:
                flat_list.append(item)
    return (flat_list)
# formats
thin = Side(border_style="thin", color="000000")
fillRed = PatternFill("solid", fgColor="FFC7CE")
redFont = Font(name='Calibri',
               size=11,
               bold=False,
               italic=False,
               vertAlign=None,
               underline='none',
               strike=False,
               color='9C0006')
fillGreen = PatternFill("solid", fgColor="1F497D")
greenFont = Font(name='Calibri',
                 size=11,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FFFFFF')
border = Border(top=thin, left=thin, right=thin, bottom=thin)
'''
top = Border(top=border.top)
left = Border(left=border.left)
right = Border(right=border.right)
bottom = Border(bottom=border.bottom)
'''

# BIM Style
def BIMStyle(cell, Fill, Font):
    cell.fill = Fill
    cell.font = Font

# CATEGORY IN WORKSETS CONTROL
def FormattingCategory(workSheets):
    wsSix = workSheets
    wsSix['A1'] = 'Model Categories'
    wsSix['B1'] = 'Worksets'
    wsSix['C1'] = 'Suggestion from BIM Team'
    wsSix['D1'] = 'Response from Project Team'
    i = 1
    # Setup Color in *Linked
    for column in wsSix['A:B']:
        for cell in column:
            i += 1
            if not cell.value is None:
                if cell.value[0] == '*':
                    BIMStyle(cell, fillRed, redFont)
                    comment = 'C' + str(cell.row)
                    wsSix[comment] = 'Model element should not be hosted in *Linked Workset'
    for cell in wsSix['C']:
        if not cell.value is None:
            BIMStyle(cell, fillRed, redFont)

    rows = wsSix['A1:D1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)
    wsSix['C1'] = 'Suggestion from BIM Team'
    '''
    wsSix.auto_filter.ref = "A1:C200"
    wsSix.auto_filter.add_filter_column(0, ["Model Categories", "Worksets", "Suggestion from BIM Team"])
    wsSix.auto_filter.add_sort_condition("C2:C200")
    '''
# Dimensions CONTROL
def FormattingDimentions(workSheet):
    ws = workSheet
    for cell in ws['B']:
        if not cell.value == 'Arial':
            BIMStyle(cell, fillRed, redFont)
            comment = 'D' + str(cell.row)
            ws[comment] = 'Please only use "Arial" as your dimension text style.'

    for cell in ws['A']:
        if str(cell.value)[0:5] != 'PA - ':
            BIMStyle(cell, fillRed, redFont)
            if ws['D' + str(cell.row)].value is None:
                ws['D' + str(cell.row)] = 'All Dim Style must start with \'PA - \' '
            else:
                ws['D' + str(cell.row)] = str(ws['D' + str(cell.row)].value) + \
                                          '; All Dim Style must start with \'PA - \' '

    for cell in ws['C'][1:]:
        if float(cell.value) > 0.12500:
            BIMStyle(cell, fillRed, redFont)
            if ws['D' + str(cell.row)].value is None:
                ws['D' + str(cell.row)] = 'Dimension tolerance should not be bigger than 1/8".'
            else:
                ws['D' + str(cell.row)] = str(ws['D' + str(cell.row)].value) + \
                                          '; Dimension tolerance should not be bigger than 1/8".'

    for cell in ws['D']:
        if not cell.value is None:
            BIMStyle(cell, fillRed, redFont)

    ws['D1'] = 'Suggestion from BIM Team'
    ws['E1'] = 'Response from Project Team'
    ws['A1'] = 'Dimension Name'
    ws['B1'] = 'Text Style'
    ws['C1'] = 'Accuracy'

    rows = ws['A1:E1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)

 # TEXT STYLE CONTROL
def FormattingText(workSheet):
    wsOne = workSheet
    textName = []
    for cell in wsOne['A']:
        if cell.value != ' ':
            textName.append(cell.value)
        else:
            break

    for cellB in wsOne['B']:
        if cellB.value != 'Arial':
            BIMStyle(cellB, fillRed, redFont)
            wsOne['M' + str(cellB.row)] = 'Please Correct the font, has to be Arial'

    for cellA in wsOne['A']:
        if str(cellA.value)[0:5] != 'PA - ':
            # TODO: Recognition of Duplicate Text Styles
            BIMStyle(cellA, fillRed, redFont)
            if wsOne['M' + str(cellA.row)].value is None:
                wsOne['M' + str(cellA.row)] = 'User created Text Style must start with \'PA - \' '
            else:
                wsOne['M' + str(cellA.row)] = wsOne['M' + str(cellA.row)].value + \
                                              '; User created Text Style must start with \'PA - \' '

    wsOne['M1'] = 'Suggestion from BIM Team'
    # Comment cell color
    for cellN in wsOne['M']:
        if not cellN.value is None:
            BIMStyle(cellN, fillRed, redFont)

    wsOne['A1'] = 'Type Name'
    wsOne['B1'] = 'Text Font'
    wsOne['C1'] = 'Text Size'
    wsOne['D1'] = 'Tab Size'
    wsOne['E1'] = 'Bold'
    wsOne['F1'] = 'Italic'
    wsOne['G1'] = 'Underline'
    wsOne['H1'] = 'Width Factor'
    wsOne['I1'] = 'Show Border'
    wsOne['J1'] = 'Background'
    wsOne['K1'] = 'Color'
    wsOne['L1'] = 'Count'
    wsOne['M1'] = 'Suggestion from BIM Team'
    wsOne['N1'] = 'Response from Project Team'
    rows = wsOne['A1:N1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)

 # LINE STYLE CONTROL
def FormattingLine(workSheet):
    ws = workSheet
    ws['G1'] = 'Suggestion from BIM Team'
    for cellA in ws['A']:
        if str(cellA.value)[0:5] != 'PA - ':
            # TODO: Recognition of Duplicate Line Styles
            BIMStyle(cellA, fillRed, redFont)
            if ws['G' + str(cellA.row)].value is None:
                ws['G' + str(cellA.row)] = 'User created Text Style must start with \'PA - \' '
            else:
                ws['G' + str(cellA.row)] = str(ws['G' + str(cellA.row)].value) + \
                                           '; User created Text Style must start with \'PA - \' '

    # Comment cell color
    for cellN in ws['G']:
        if not cellN.value is None:
            BIMStyle(cellN, fillRed, redFont)

    ws['A1'] = 'Line Type'
    ws['B1'] = 'Red'
    ws['C1'] = 'Green'
    ws['D1'] = 'Blue'
    ws['E1'] = 'Line Weight'
    ws['F1'] = 'Style'
    ws['G1'] = 'Suggestion from BIM Team'
    ws['H1'] = 'Response from Project Team'
    rows = ws['A1:H1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)

def FormattingSettings(workSheet, phases):
    ws = workSheet
    cellA = ws['B2']
    if str(cellA) > '0.126':
        BIMStyle(cellA, fillRed, redFont)
        if ws['D' + str(cellA.row)].value is None:
            ws['D' + str(cellA.row)] = 'Project Tolerance too big '
        else:
            ws['D' + str(cellA.row)] = str(ws['D' + str(cellA.row)].value) + \
                                       '; Project Tolerance too big '
    for cell in ws['3']:
        if cell.value not in phases:
            BIMStyle(cell, fillRed, redFont)
    # Comment cell color
    for cellN in ws['D']:
        if not cellN.value is None:
            BIMStyle(cellN, fillRed, redFont)

    ws['A1'] = 'Project Settings'
    ws['D1'] = 'Suggestion from BIM Team'
    ws['E1'] = 'Response from Project Team'
    rows = ws['A1:E1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)

def FormattingFilledRegion(workSheet):
    ws = workSheet
    ws['B1'] = 'Suggestion from BIM Team'
    for cellA in ws['A']:
        if str(cellA.value)[0:5] != 'PA - ':
            # TODO: Recognition of Duplicate Region Styles
            BIMStyle(cellA, fillRed, redFont)
            if ws['B' + str(cellA.row)].value is None:
                ws['B' + str(cellA.row)] = 'All Filled Region must start with \'PA - \' '
            else:
                ws['B' + str(cellA.row)] = str(ws['B' + str(cellA.row)].value) + \
                                           '; All Filled Region must start with \'PA - \' '

    # Comment cell color
    for cellN in ws['B']:
        if not cellN.value is None:
            BIMStyle(cellN, fillRed, redFont)

    ws['A1'] = 'Filled Region'
    ws['B1'] = 'Suggestion from BIM Team'
    ws['C1'] = 'Response from Project Team'
    rows = ws['A1:C1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)

def FormattingLevel(workSheet, fileDis):
    ws = workSheet
    ws['B1'] = 'Suggestion from BIM Team'
    for cellA in ws['A']:
        if str(cellA.value)[0] != fileDis + ' - ':
            # TODO: Recognition of Duplicate Region Styles
            BIMStyle(cellA, fillRed, redFont)
            if ws['B' + str(cellA.row)].value is None:
                ws['B' + str(cellA.row)] = 'Levels should start with Discipline, for example \'A - Arrivals\' '
            else:
                ws['B' + str(cellA.row)] = str(ws['B' + str(cellA.row)].value) + \
                                           '; Levels should start with Discipline, for example \'A - Arrivals\' '

    # Comment cell color
    for cellN in ws['B']:
        if not cellN.value is None:
            BIMStyle(cellN, fillRed, redFont)

    ws['A1'] = 'Level'
    ws['B1'] = 'Suggestion from BIM Team'
    ws['C1'] = 'Response from Project Team'
    rows = ws['A1:C1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)

def FormattingAnnotationSymbol(workSheet):
    ws = workSheet
    ws['B1'] = 'Suggestion from BIM Team'
    for cellA in ws['A']:
        if str(cellA.value)[0:5] != 'PA - ':
            BIMStyle(cellA, fillRed, redFont)
            if ws['B' + str(cellA.row)].value is None:
                ws['B' + str(cellA.row)] = 'All Filled Region must start with \'PA - \' '
            else:
                ws['B' + str(cellA.row)] = str(ws['B' + str(cellA.row)].value) + \
                                           '; All Filled Region must start with \'PA - \' '

    # Comment cell color
    for cellN in ws['B']:
        if not cellN.value is None:
            BIMStyle(cellN, fillRed, redFont)

    ws['A1'] = 'Annotation Symbol'
    ws['B1'] = 'Suggestion from BIM Team'
    ws['C1'] = 'Response from Project Team'
    rows = ws['A1:C1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)

def FormattingWorkset(worksheet):
    ws = worksheet
    for row in ws.rows:
        for cell in row:
            if cell.value == 'Model Coordination' and ws['B' + str(cell.row)] == 'True':
                BIMStyle(cell, fillRed, redFont)
                ws['C' + str(cell.row)] = 'Model Coordination needs to be hidden by default'
            elif str(cell.value)[0] == '*' and str(cell.value)[1:5] != 'LINK':
                BIMStyle(cell, fillRed, redFont)
                ws['C' + str(cell.row)] = 'Normal workset should not have "*" in front'
    for cellA in ws['A']:
        if not '-' in str(cellA.value) and cellA.value != 'Model Coordination':
            BIMStyle(cellA, fillRed, redFont)
            if ws['C' + str(cellA.row)].value is None:
                ws['C' + str(cellA.row)] = 'Workset naming not to PA Standard, ' \
                                           'should be XXXXX-XXXX ' \
                                           'example: Head House-Ceiling'
            else:
                ws['C' + str(cellA.row)] = str(ws['B' + str(cellA.row)].value) + '; Workset naming not to PA Standard, ' \
                                                                                 'should be XXXXX-XXXX ' \
                                                                                 'example: Head House-Ceiling'

    for cell in ws['C']:
        if not cell.value is None:
            BIMStyle(cell, fillRed, redFont)

    ws['A1'] = 'Worksets'
    ws['B1'] = 'Visible by Default'
    ws['C1'] = 'Suggestion from BIM Team'
    ws['D1'] = 'Response from Project Team'
    rows = ws['A1:D1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)

    # Sheets CONTROL
def FormattingSheets(workSheet, viewClassification, discipline, sub1, sub2, sub3):
    wsEight = workSheet
    wsEight['A1'] = 'Sheet Number'
    wsEight['B1'] = 'Sheet Name'
    wsEight['C1'] = 'PA - View Classification'
    wsEight['D1'] = 'Appear in sheet list'
    wsEight['E1'] = 'Discipline Group'
    wsEight['F1'] = 'Discipline Sub-Group'
    wsEight['G1'] = 'Suggestion from BIM Team'
    wsEight['H1'] = 'Response from Project Team'
    sheetRegex = re.compile(r'^\w\w?w?\d\d\D\d\d\d$')
    i = 1
    # Setup Color in *Linked
    for cell in wsEight['C']:
        i += 1
        appear = 'D' + str(cell.row)
        if wsEight[appear].value == True:
            if cell.value == 'No Value':
                BIMStyle(cell, fillRed, redFont)
                comment = 'G' + str(cell.row)
                wsEight[comment] = 'PA - View Classification parameter should not be empty'
            elif not cell.value in viewClassification:
                BIMStyle(cell, fillRed, redFont)
                comment = 'G' + str(cell.row)
                wsEight[comment] = 'PA - View Classification parameter should follow ' \
                               'PA - View Classification Standard on PA BIM Standard page 87'

    for cell in wsEight['A']:
        appear = 'D' + str(cell.row)
        if wsEight[appear].value == True:
            if sheetRegex.findall(str(cell.value)) == []:
                BIMStyle(cell, fillRed, redFont)
                comment = 'G' + str(cell.row)
                if wsEight[comment].value is None:
                    wsEight[comment] = 'PA - Sheet Number incorrect please see BEP for Sheet Numbering Convention'
                else:
                    wsEight[comment] = wsEight[comment].value + '; PA - Sheet Number incorrect please see BEP ' \
                                                            'for Sheet Numbering Convention, if no level info, please '
            elif not cell.value[0].isupper():
                BIMStyle(cell, fillRed, redFont)
                comment = 'G' + str(cell.row)
                if wsEight[comment].value is None:
                    wsEight[comment] = 'First Discipline Code should be uppercase, ' \
                                       'please confirm if this is indeed a Discipline letter or an error.'
                else:
                    wsEight[comment] = wsEight[comment].value + '; First Discipline Code should be uppercase,' \
                                                                'please confirm if this is indeed a ' \
                                                                'Discipline letter or an error.'
    for cell in wsEight['E']:

        if cell.value == 'No Value':
            BIMStyle(cell, fillRed, redFont)
            comment = 'G' + str(cell.row)
            if wsEight[comment].value is None:
                wsEight[comment] = 'Discipline parameter needs to be filled. '
            else:
                wsEight[comment] = wsEight[comment].value + '; Discipline parameter needs to be filled'

        elif not cell.value in discipline:
            comment = 'G' + str(cell.row)
            if wsEight[comment].value is None:
                wsEight[comment] = 'Discipline parameter not correct '
            else:
                wsEight[comment] = wsEight[comment].value + '; Discipline parameter not correct '
        elif cell.value == discipline[0]:
            subGroup = wsEight['F' + str(cell.row)].value
            if subGroup not in sub1:
                BIMStyle(cell, fillRed, redFont)
                comment = 'G' + str(cell.row)
                if wsEight[comment].value is None:
                    wsEight[comment] = 'Sub-Discipline not standard, please check P 164 of PA Standard'
                else:
                    wsEight[comment] = wsEight[comment].value + '; Sub-Discipline not standard, please check P 164 of PA Standard'
        elif cell.value == discipline[1]:
            subGroup = wsEight['F' + str(cell.row)].value
            if subGroup not in sub2:
                BIMStyle(cell, fillRed, redFont)
                comment = 'G' + str(cell.row)
                if wsEight[comment].value is None:
                    wsEight[comment] = 'Sub-Discipline not standard, please check P 164 of PA Standard'
                else:
                    wsEight[comment] = wsEight[comment].value + '; Sub-Discipline not standard, please check P 164 of PA Standard'
        elif cell.value == discipline[2]:
            subGroup = wsEight['F' + str(cell.row)].value
            if subGroup not in sub3:
                BIMStyle(cell, fillRed, redFont)
                comment = 'G' + str(cell.row)
                if wsEight[comment].value is None:
                    wsEight[comment] = 'Sub-Discipline not standard, please check P 164 of PA Standard'
                else:
                    wsEight[comment] = wsEight[comment].value + '; Sub-Discipline not standard, please check P 164 of PA Standard'

    # TODO: Change these
    for cell in wsEight['E']:
        appear = 'E' + str(cell.row)
        comment = 'F' + str(cell.row)
        if wsEight[comment].value is None:
            wsEight[comment] = ''
        else:
            wsEight[comment] = wsEight[comment].value + ''

    for cell in wsEight['G']:
        if not cell.value is None:
            BIMStyle(cell, fillRed, redFont)

    rows = wsEight['A1:H1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)
    wsEight['G1'] = 'Suggestion from BIM Team'


    # PROJECT INFO CONTROL
def FormattingProjectInfo(workSheet):
    wsThree = workSheet
    wsThree['A1'] = 'Points'
    wsThree['A2'] = 'Project Base Point'
    wsThree['A3'] = 'Project Survey Point'
    wsThree['A4'] = 'Shared Point'
    wsThree['B1'] = 'X'
    wsThree['C1'] = 'Y'
    wsThree['D1'] = 'Z'
    wsThree['E1'] = 'Angle'
    wsThree['F1'] = 'Pinned'
    wsThree['G1'] = 'Discipline'
    wsThree['H1'] = 'Workset'
    wsThree['I1'] = 'Suggestion from BIM Team'
    wsThree['J1'] = 'Response from Project Team'
    if wsThree['C2'].value != 674500 or \
            wsThree['B2'].value != 579600 or \
            wsThree['D2'].value != 0 or \
            wsThree['E2'].value != 94.75:
        BIMStyle(wsThree['A2'], fillRed, redFont)
        wsThree['I2'] = 'Location issue for Project Base Point.'
    if wsThree['C3'].value != 0 or \
            wsThree['B3'].value != 0 or \
            wsThree['D3'].value != 0:
        BIMStyle(wsThree['A3'], fillRed, redFont)
        wsThree['I3'] = 'Location issue for Project Survey Point.'
    try :
        if float(wsThree['B4'].value) < 578564.9 or float(wsThree['B4'].value) > 578565.0 or \
                float(wsThree['C4'].value) < 674225.9 or float(wsThree['C4'].value) > 674226.0 or\
                float(wsThree['D4'].value) < 10.9999 or float(wsThree['D4'].value) > 11.001:
            BIMStyle(wsThree['A4'], fillRed, redFont)
            wsThree['I4'] = 'Location issue for Project Shared Point.'
    except:
        pass

    if not wsThree['F4'].value == 'Pinned' or wsThree['H4'] == 'Workset : Model Coordination':
        BIMStyle(wsThree['A4'], fillRed, redFont)
        if wsThree['I4'].value is None:
            wsThree['I4'] = 'Pin or Workset issue for Shared Point. '
        else:
            wsThree['I4'] = wsThree['I4'].value + '; Pin or Workset issue for Shared Point.'
    for cell in wsThree['I']:
        if not cell.value is None:
            BIMStyle(cell, fillRed, redFont)
    rows = wsThree['A1:J1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)
    wsThree['I1'] = 'Suggestion from BIM Team'

def FormattingFamily(workSheet):
    # Family Name CONTROL
    ws = workSheet
    ws['A1'] = 'Category'
    ws['B1'] = 'Family Name'
    ws['C1'] = 'Suggestion from BIM Team'
    ws['D1'] = 'Response from Project Team'
    familyRegex = re.compile(r'\S+.*\s?-\s?\S+.*\s?-\s?\S+.*?')
    for cell in ws['A']:
        if cell.value == 'Generic Models':
            BIMStyle(cell, fillRed, redFont)
            comment = 'C' + str(cell.row)
            if ws[comment].value is None:
                ws[comment] = 'Please consider changing Generic Models to your discipline category.'
            else:
                ws[comment] = ws[comment].value + '; Please consider changing Generic ' \
                                                  'Models to your discipline category.'
    for cell in ws['B']:
        ccate = 'A' + str(cell.row)
        tcate = str(ws[ccate].value)
        # need a form to include the singular of the category name
        tcateMs = tcate[0: len(tcate) - 1]
        cateRegex = re.compile(r'^' + tcate + r'(.*)')
        cateMsRegex = re.compile(r'^' + tcateMs + r'(.*)')
        x = len(tcate)

        if not cell.value in Importcsv('APFamily.csv'):
            if familyRegex.findall(cell.value) == [] and cell.value[0:2] != 'PA':
                BIMStyle(cell, fillRed, redFont)
                comment = 'C' + str(cell.row)
                ws[comment] = 'Please rename family to <Category>-<Manufacturer>-<Description and/or ' \
                                  'Model Number> to follow PA Family naming Convention on PA BIM Standard Page 62-63'
            elif cell.value[0:x] != tcate and cell.value[0:x - 1] != tcateMs and cell.value[0:2] != 'PA':
                BIMStyle(cell, fillRed, redFont)
                comment = 'C' + str(cell.row)
                ws[comment] = 'Family name needs to start with <Category>, ' \
                                  'please follow PA Family naming Convention on PA BIM Standard Page 62-63'
            elif cell.value[0:2] == 'PA':
                BIMStyle(cell, fillRed, redFont)
                comment = 'C' + str(cell.row)
                ws[comment] = 'Please make sure this family came from Port Authority, ' \
                                  'if created by you or you colleague, it should not be named "PA"'

    for cell in ws['C']:
        if not cell.value is None:
            BIMStyle(cell, fillRed, redFont)

    rows = ws['A1:D1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)
    ws['C1'] = 'Suggestion from BIM Team'
    ws.auto_filter.ref = "A1:C2000"
    ws.auto_filter.add_filter_column(0, ["Category", "Worksets", "Suggestion from BIM Team"])
    ws.auto_filter.add_sort_condition("C2:C2000")


def FormattingSharedParameters(workSheet, wbSharedParameter):
    wsTen = workSheet
    wsTen['A1'] = 'Parameter Name'
    wsTen['B1'] = 'Element ID'
    wsTen['C1'] = 'GUID'
    wsTen['D1'] = 'Suggestion from BIM Team'
    wsSP = wbSharedParameter['PA - Shared_Parameters']
    SharedParameterID = []
    SharedParameter = []
    MisLstID = []
    MisLst = []
    p = 1
    for i in wsTen['C']:
        SharedParameterID.append(i.value)
        sName = wsTen['A' + str(p)]
        SharedParameter.append(sName.value)
        p += 1

    o = 1
    for cell in wsSP['B']:
            if not cell.value in SharedParameterID:
                MisLstID.append(cell.value)
                oName = wsSP['C' + str(o)]
                MisLst.append(oName.value)
                o += 1
    k = 1
    for cell in wsTen['B']:
        k += 1
    m = 0
    for val in MisLstID:
        cell = wsTen['C' + str(k)]
        wsTen['C' + str(k)].value = val
        kName = MisLst[m]
        wsTen['A' + str(k)].value = kName
        BIMStyle(cell, fillRed, redFont)
        k += 1
        m += 1

    rows = wsTen['A1:D1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)
    wsTen['D1'] = 'Suggestion from BIM Team'

def FormattingTitleBlock(workSheet):
    # TITLE BLOCK CONTROL
    wsTwo = workSheet
    wsTwo['A1'] = 'Sheet Number'
    wsTwo['B1'] = 'Family Name'
    wsTwo['C1'] = 'Suggestion from BIM Team'
    wsTwo['D1'] = 'Response from Project Team'
    for cell in wsTwo['B']:
        if str(cell.value)[0] != 'P' or str(cell.value)[1] != 'A':
            BIMStyle(cell, fillRed, redFont)
            wsTwo['C' + str(cell.row)] = 'Please use PA Approved Titleblock'
    # Comment cell color
    for cell in wsTwo['C']:
        if not cell.value is None:
            BIMStyle(cell, fillRed, redFont)

    rows = wsTwo['A1:D1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)
    wsTwo['C1'] = 'Suggestion from BIM Team'

def FormattingViews(workSheet, dCoderaw, viewClassification):

    wsSeven = workSheet
    wsSeven['A1'] = 'View Name'
    wsSeven['B1'] = 'PA - View Classification'
    wsSeven['C1'] = 'Suggestion from BIM Team'
    wsSeven['D1'] = 'Response from Project Team'
    viewRegex = re.compile(r'^\w\w-\S\S-(.*)')
    viewRegex02 = re.compile(r'^\w\w\S\S-(.*)')

    i = 1

    dCodes = []
    for pp in dCoderaw:
        dCodes.append(pp[0:2])

    # Setup Color in *Linked

    for cell in wsSeven['B']:

        i += 1
        if cell.value is None:
            BIMStyle(cell, fillRed, redFont)
            comment = 'C' + str(cell.row)
            wsSeven[comment] = 'PA - View Classification parameter should not be empty'
        elif not cell.value in viewClassification:
            BIMStyle(cell, fillRed, redFont)
            comment = 'C' + str(cell.row)
            wsSeven[
                comment] = 'PA - View Classification parameter should follow ' \
                           'PA - View Classification Standard on PA BIM Standard page 87'
    for cell in wsSeven['A']:
        if viewRegex.findall(cell.value) == [] and viewRegex02.findall(cell.value) == []:
            BIMStyle(cell, fillRed, redFont)
            comment = 'C' + str(cell.row)
            if wsSeven[comment].value is None:
                wsSeven[comment] = 'View Naming incorrect please see BEP for View Naming Convention'
            else:
                wsSeven[comment] = wsSeven[comment].value + '; View Naming incorrect please see ' \
                                                            'BEP for View Naming Convention'
        elif viewRegex02.findall(str(cell.value)) != []:
            BIMStyle(cell, fillRed, redFont)
            comment = 'C' + str(cell.row)
            if wsSeven[comment].value is None:
                wsSeven[comment] = 'should have "-" between Type code and level, For example"FP01" should be "FP-01"'
            else:
                wsSeven[comment] = wsSeven[
                                       comment].value + '; should have "-" between Type code and level, ' \
                                                        'For example"FP01" should be "FP-01"'
        elif len(str(cell.value)) > 30:
            BIMStyle(cell, fillRed, redFont)
            comment = 'C' + str(cell.row)
            if wsSeven[comment].value is None:
                wsSeven[
                    comment] = 'Description too long, should not need level information in Description, ' \
                               'please limit the description to less than 24 charaters'
            else:
                wsSeven[comment] = wsSeven[comment].value + '; Description too long, ' \
                                                        'should not need level information in Description'
        elif not str(cell.value)[0:2] in dCodes and str(cell.value)[0] != 'L':
            BIMStyle(cell, fillRed, redFont)
            comment = 'C' + str(cell.row)
            if wsSeven[comment].value is None:
                wsSeven[comment] = 'View Type Code not permitted, please see PA BIM Standard page 81-84 for information'
            else:
                wsSeven[comment] = wsSeven[comment].value + '; View Type Code not permitted, ' \
                                                        'please see PA BIM Standard page 81-84 for information'

    for cell in wsSeven['C']:
        if not cell.value is None:
            BIMStyle(cell, fillRed, redFont)

    rows = wsSeven['A1:D1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)
    wsSeven['C1'] = 'Suggestion from BIM Team'

def FormattingLinks(workSheet):
    # LINKS CONTROL
    wsFour = workSheet
    wsFour['A1'] = 'Link Name'
    wsFour['B1'] = 'Location'
    wsFour['C1'] = 'Workset'
    wsFour['D1'] = 'Pinned or not'
    wsFour['E1'] = 'Attachment Type'
    wsFour['F1'] = 'Suggestion from BIM Team'
    wsFour['G1'] = 'Response from Project Team'
    for cell in wsFour['B']:
        if cell.value != ' location <Not Shared>':
            BIMStyle(cell, fillRed, redFont)
            wsFour['F' + str(cell.row)] = 'Link Location needs to be \'Not Shared\''
    for cell in wsFour['C']:
        linkname = wsFour['A' + str(cell.row)]
        if linkname.value == 'X17017000-SCOPEBOXES.rvt ':
            BIMStyle(linkname, fillRed, redFont)
            comment = 'F' + str(cell.row)
            if wsFour[comment].value is None:
                wsFour[comment] = 'Scopeboxes file does not need to be kept loaded into the model, ' \
                                  'please remove this link '
            else:
                wsFour[comment] = wsFour[comment].value + '; Scopeboxes file does not need to be kept ' \
                                                          'loaded into the model, please remove this link '
        elif cell.value[8:17] != linkname.value[0:9]:
            BIMStyle(cell, fillRed, redFont)
            comment = 'F' + str(cell.row)
            if wsFour[comment].value is None:
                wsFour[comment] = 'Each Linked Model needs to be in a properly named workset in BEP. '
            else:
                wsFour[comment] = wsFour[comment].value + '; Each Linked Model needs to be ' \
                                                          'in a properly named workset in BEP. '
    for cell in wsFour['D']:
        if cell.value == 'Not Pinned':
            BIMStyle(cell, fillRed, redFont)
            comment = 'F' + str(cell.row)
            if wsFour[comment].value is None:
                wsFour[comment] = 'Linked Model needs to be pinned. '
            else:
                wsFour[comment] = wsFour[comment].value + '; Linked Model needs to be pinned. '
    for cell in wsFour['E']:
        if cell.value =='Attachment':
            BIMStyle(cell, fillRed, redFont)
            comment = 'F' + str(cell.row)
            if wsFour[comment].value is None:
                wsFour[comment] = 'Linked Type needs to be Overlay. '
            else:
                wsFour[comment] = wsFour[comment].value + '; Linked Type needs to be Overlay. '
    # Format Comment
    for cell in wsFour['F']:
        if not cell.value is None:
            BIMStyle(cell, fillRed, redFont)

    wsFour['F1'] = 'Suggestion from BIM Team'
    rows = wsFour['A1:G1']
    for cell in rows[0]:
        BIMStyle(cell, fillGreen, greenFont)